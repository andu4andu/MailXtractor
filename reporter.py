import logging
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_MAX_COL_WIDTH = 50


def _style_sheet(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[letter].width = min(width + 2, _MAX_COL_WIDTH)

    for cell in ws[1]:
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def generate_excel_report(records: list, output_path: str) -> None:
    logger.info("Generating report with %d records", len(records))

    successful = [{k: v for k, v in r.items() if k != "_error"} for r in records if not r.get("_error")]
    failed = [r for r in records if r.get("_error")]

    df_success = pd.DataFrame(successful)
    df_failed  = pd.DataFrame(failed)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_success.to_excel(writer, sheet_name="Results",  index=False)
        df_failed.to_excel(writer,  sheet_name="Failures", index=False)
        _style_sheet(writer.sheets["Results"])
        if not df_failed.empty:
            _style_sheet(writer.sheets["Failures"])

    logger.info("Report saved to %s", output_path)
    logger.info("%d successful, %d failed", len(successful), len(failed))
